import io
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Dict, Generator, List, Tuple

from genson import SchemaBuilder
from rest_framework import serializers

from baserow.contrib.integrations.core.service_types import CoreServiceType
from baserow.core.code_runner.exceptions import (
    CodeRunnerExecutionError,
    CodeRunnerImproperlyConfigured,
    CodeRunnerResultError,
)
from baserow.core.code_runner.registries import (
    get_code_runner,
)
from baserow.core.formula.types import BaserowFormulaObject
from baserow.core.formula.validator import (
    ensure_file,
    ensure_json_serializable,
    ensure_string,
)
from baserow.core.registry import Instance
from baserow.core.services.dispatch_context import DispatchContext
from baserow.core.services.exceptions import (
    ServiceImproperlyConfiguredDispatchException,
    UnexpectedDispatchException,
)
from baserow.core.services.models import Service
from baserow.core.services.registries import DispatchTypes, ListServiceTypeMixin
from baserow.core.services.types import DispatchResult, FormulaToResolve, ServiceDict
from baserow_enterprise.features import XLS_FILE_READER
from baserow_enterprise.integrations.core.models import (
    CORE_CODE_SERVICE_CODE_MAX_LENGTH,
    CoreCodeService,
    CoreCodeServiceInjection,
    CoreXLSFileReaderService,
)
from baserow_premium.license.handler import LicenseHandler


class CoreCodeServiceType(CoreServiceType):
    type = "code"
    model_class = CoreCodeService
    dispatch_types = [DispatchTypes.ACTION]
    allowed_fields = ["code"]
    serializer_field_names = ["code", "injections"]
    request_serializer_field_names = ["code", "injections"]

    class SerializedDict(ServiceDict):
        code: str
        injections: List[Dict[str, str]]

    def get_schema_name(self, service: CoreCodeService) -> str:
        return f"Code{service.id}Schema"

    def generate_schema(
        self,
        service: CoreCodeService,
        allowed_fields: List[str] | None = None,
    ) -> Dict[str, Any] | None:
        if service.sample_data is None or "data" not in service.sample_data:
            return None

        schema_builder = SchemaBuilder()
        schema_builder.add_object(service.sample_data["data"])

        return {
            **schema_builder.to_schema(),
            "title": self.get_schema_name(service),
        }

    @property
    def serializer_field_overrides(self):
        from baserow_enterprise.integrations.core.api.serializers import (
            CoreCodeServiceInjectionSerializer,
        )

        return {
            "code": serializers.CharField(
                help_text=CoreCodeService._meta.get_field("code").help_text,
                allow_blank=True,
                max_length=CORE_CODE_SERVICE_CODE_MAX_LENGTH,
                required=False,
            ),
            "injections": CoreCodeServiceInjectionSerializer(
                many=True,
                required=False,
                help_text="The values to inject into the code executor context.",
            ),
        }

    @property
    def request_serializer_field_overrides(self):
        return self.serializer_field_overrides

    def after_create(
        self,
        instance: CoreCodeService,
        values: Dict,
    ):
        if "injections" in values:
            instance.injections.all().delete()
            CoreCodeServiceInjection.objects.bulk_create(
                [
                    CoreCodeServiceInjection(
                        service=instance,
                        name=injection["name"],
                        formula=injection["formula"],
                    )
                    for injection in values["injections"]
                ]
            )

    def after_update(
        self,
        instance: CoreCodeService,
        values: Dict,
        changes: Dict[str, Tuple],
    ):
        return self.after_create(instance, values)

    def formulas_to_resolve(self, service: CoreCodeService) -> list[FormulaToResolve]:
        return [
            FormulaToResolve(
                injection.name,
                injection.formula,
                ensure_json_serializable,
                f'injection "{injection.name}"',
            )
            for injection in service.injections.all()
        ]

    def extract_properties(
        self, service: Service, path: List[str], **kwargs
    ) -> List[str]:
        if path:
            return [path[0]]
        return []

    def dispatch_data(
        self,
        service: CoreCodeService,
        resolved_values: Dict[str, Any],
        dispatch_context: DispatchContext,
    ) -> Dict[str, Any]:
        try:
            return get_code_runner().run(resolved_values, service.code)
        except CodeRunnerImproperlyConfigured as exc:
            raise ServiceImproperlyConfiguredDispatchException(str(exc)) from exc
        except CodeRunnerResultError as exc:
            raise ServiceImproperlyConfiguredDispatchException(str(exc)) from exc
        except CodeRunnerExecutionError as exc:
            raise UnexpectedDispatchException(str(exc)) from exc

    def dispatch_transform(self, data: Dict[str, Any]) -> DispatchResult:
        return DispatchResult(data=data)

    def formula_generator(
        self, service: Service
    ) -> Generator[str | Instance, str, None]:
        yield from super().formula_generator(service)

        for injection in service.injections.all():
            new_formula = yield BaserowFormulaObject.to_formula(injection.formula)
            if new_formula is not None:
                injection.formula = new_formula
                yield injection

    def serialize_property(
        self,
        service: CoreCodeService,
        prop_name: str,
        files_zip=None,
        storage=None,
        cache=None,
    ):
        if prop_name == "injections":
            return [
                {
                    "name": injection.name,
                    "formula": injection.formula,
                }
                for injection in service.injections.all()
            ]

        return super().serialize_property(
            service, prop_name, files_zip=files_zip, storage=storage, cache=cache
        )

    def create_instance_from_serialized(
        self,
        serialized_values,
        id_mapping,
        files_zip=None,
        storage=None,
        cache=None,
        **kwargs,
    ):
        injections = serialized_values.pop("injections", [])

        service = super().create_instance_from_serialized(
            serialized_values,
            id_mapping,
            files_zip=files_zip,
            storage=storage,
            cache=cache,
            **kwargs,
        )

        CoreCodeServiceInjection.objects.bulk_create(
            [
                CoreCodeServiceInjection(
                    **injection,
                    service=service,
                )
                for injection in injections
            ]
        )

        return service

    def enhance_queryset(self, queryset):
        return super().enhance_queryset(queryset).prefetch_related("injections")


class CoreXLSFileReaderServiceType(ListServiceTypeMixin, CoreServiceType):
    type = "xls_file_reader"
    model_class = CoreXLSFileReaderService
    dispatch_types = [DispatchTypes.DATA, DispatchTypes.ACTION]

    allowed_fields = [
        "file",
        "sheet_name",
        "first_line_is_header",
    ]

    serializer_field_names = [
        "file",
        "sheet_name",
        "first_line_is_header",
    ]

    class SerializedDict(ServiceDict):
        file: str
        sheet_name: str
        first_line_is_header: bool

    simple_formula_fields = [
        "file",
        "sheet_name",
    ]

    def is_deactivated(self, workspace) -> bool:
        return not LicenseHandler.workspace_has_feature(XLS_FILE_READER, workspace)

    def raise_if_deactivated(self, workspace) -> None:
        LicenseHandler.raise_if_workspace_doesnt_have_feature(
            XLS_FILE_READER, workspace
        )

    def dispatch(
        self,
        service: CoreXLSFileReaderService,
        dispatch_context: DispatchContext,
    ) -> DispatchResult:
        self.raise_if_deactivated(dispatch_context.workspace)

        return super().dispatch(service, dispatch_context)

    @property
    def serializer_field_overrides(self):
        from baserow.core.formula.serializers import FormulaSerializerField

        return {
            "file": FormulaSerializerField(
                help_text=CoreXLSFileReaderService._meta.get_field("file").help_text,
                required=False,
            ),
            "sheet_name": FormulaSerializerField(
                help_text=CoreXLSFileReaderService._meta.get_field(
                    "sheet_name"
                ).help_text,
                required=False,
            ),
            "first_line_is_header": serializers.BooleanField(
                help_text=CoreXLSFileReaderService._meta.get_field(
                    "first_line_is_header"
                ).help_text,
                required=False,
            ),
        }

    def get_schema_name(self, service: CoreXLSFileReaderService) -> str:
        return f"XLSFileReader{service.id}Schema"

    def generate_schema(
        self,
        service: CoreXLSFileReaderService,
        allowed_fields: List[str] | None = None,
    ) -> Dict[str, Any] | None:
        if (
            service.sample_data
            and "data" in service.sample_data
            and "results" in service.sample_data["data"]
            and (allowed_fields is None or "items" in allowed_fields)
        ):
            schema_builder = SchemaBuilder()
            schema_builder.add_object(service.sample_data["data"]["results"])
            schema = schema_builder.to_schema()

            if "items" in schema:
                return {
                    **schema,
                    "title": self.get_schema_name(service),
                }

        return None

    def formulas_to_resolve(
        self, service: CoreXLSFileReaderService
    ) -> list[FormulaToResolve]:
        formulas = [
            FormulaToResolve(
                "file",
                service.file,
                ensure_file,
                "'file' property",
            )
        ]

        if service.sheet_name:
            formulas.append(
                FormulaToResolve(
                    "sheet_name",
                    service.sheet_name,
                    ensure_string,
                    "'sheet_name' property",
                )
            )

        return formulas

    def extract_properties(
        self, service: Service, path: List[str], **kwargs
    ) -> List[str]:
        if path:
            return [path[0]]
        return []

    def dispatch_data(
        self,
        service: CoreXLSFileReaderService,
        resolved_values: Dict[str, Any],
        dispatch_context: DispatchContext,
    ) -> Any:
        rows = self._read_xls_file(
            resolved_values["file"],
            resolved_values.get("sheet_name", ""),
        )

        return {
            "results": self._rows_to_dicts(rows, service.first_line_is_header),
            "has_next_page": False,
        }

    def dispatch_transform(
        self,
        data: Any,
    ) -> DispatchResult:
        return DispatchResult(data=data)

    def _read_xls_file(self, input_file, sheet_name: str) -> list[list[str]]:
        data = input_file.read_bytes()
        if self._is_legacy_xls(data):
            return self._read_legacy_xls(data, sheet_name)
        return self._read_openxml_workbook(data, sheet_name)

    def _read_openxml_workbook(self, data: bytes, sheet_name: str) -> list[list[str]]:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        try:
            workbook = load_workbook(
                io.BytesIO(data),
                read_only=True,
                data_only=True,
            )
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ServiceImproperlyConfiguredDispatchException(
                f"The XLSX file couldn't be read: {exc}."
            ) from exc

        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ServiceImproperlyConfiguredDispatchException(
                        f'The sheet "{sheet_name}" does not exist.'
                    )
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.worksheets[0]

            return [
                [self._cell_value_to_string(value) for value in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()

    def _read_legacy_xls(self, data: bytes, sheet_name: str) -> list[list[str]]:
        from xlrd import XLRDError, open_workbook
        from xlrd.compdoc import CompDocError

        try:
            workbook = open_workbook(file_contents=data)
        except (CompDocError, XLRDError) as exc:
            raise ServiceImproperlyConfiguredDispatchException(
                f"The XLS file couldn't be read: {exc}."
            ) from exc

        if workbook.nsheets == 0:
            return []

        try:
            sheet = (
                workbook.sheet_by_name(sheet_name)
                if sheet_name
                else workbook.sheet_by_index(0)
            )
        except XLRDError as exc:
            if sheet_name:
                raise ServiceImproperlyConfiguredDispatchException(
                    f'The sheet "{sheet_name}" does not exist.'
                ) from exc
            raise ServiceImproperlyConfiguredDispatchException(
                f"The XLS file couldn't be read: {exc}."
            ) from exc

        return [
            [
                self._cell_value_to_string(sheet.cell_value(rowx, colx))
                for colx in range(sheet.ncols)
            ]
            for rowx in range(sheet.nrows)
        ]

    def _rows_to_dicts(
        self,
        rows: list[list[str]],
        first_line_is_header: bool,
    ) -> list[dict[str, str]]:
        if not rows:
            return []

        if first_line_is_header:
            headers = [header.strip() for header in rows[0]]
            return [
                self._row_to_dict(headers, row)
                for row in rows[1:]
                if self._row_has_values(row)
            ]

        headers = [f"column_{index + 1}" for index in range(self._widest_row(rows))]
        return [
            self._row_to_dict(headers, row) for row in rows if self._row_has_values(row)
        ]

    def _row_to_dict(self, headers: list[str], row: list[str]) -> dict[str, str]:
        return {
            header or f"column_{index + 1}": row[index] if index < len(row) else ""
            for index, header in enumerate(headers)
        }

    def _cell_value_to_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date | time):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, Decimal):
            return str(value)
        return str(value)

    def _row_has_values(self, row: list[str]) -> bool:
        return any(value != "" for value in row)

    def _widest_row(self, rows: list[list[str]]) -> int:
        return max((len(row) for row in rows), default=0)

    def _is_legacy_xls(self, data: bytes) -> bool:
        return data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")
